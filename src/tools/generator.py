import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

def destroyInfo():
    ruta_json = "../output/data.json"

    try:
        with open(ruta_json, 'w', encoding='utf-8') as archivo:
            json.dump({}, archivo, indent=4)
        print("El contenido de data.json ha sido vaciado correctamente.")
    except Exception as e:
        print(f"Error al vaciar el archivo JSON: {e}")

def generate():
    output_file = "salida.xlsx"
    with open("../output/data.json", "r", encoding="utf-8") as file:
        data = json.load(file)

    general_info = data["general"]
    estudiantes = data["ESTUDIANTES"]

    # Crear una lista para almacenar los datos procesados
    data_rows = []

    # Recorrer cada estudiante y sus materias
    for estudiante in estudiantes:
        matricula = estudiante["MATRICULA"]
        for materia in estudiante["MATERIAS"]:
            data_rows.append([
                matricula,
                materia["CDEF"],
                general_info["CODIGO_PERIODO_ESCOLAR"],
                materia["CLAVE_MATERIA"],
                general_info["FECHA_ASIGNACION"],
                general_info["CUATRIMESTRE_CURSADO"],
                general_info["NIVEL_EDUCATIVO"],
                general_info["PROFESOR_NUMERO_CONTROL"],
                general_info["TIPO_EVALUACION"]
            ])

    # Crear un DataFrame de Pandas
    columns = [
        "Matricula", "Calificacion", "Codigo Periodo", "Clave Materia", "Fecha Asignacion",
        "Cuatrimestre Cursado", "Nivel Educativo", "Profesor Numero Control", "Tipo Evaluacion"
    ]
    df = pd.DataFrame(data_rows, columns=columns)

    # Guardar en un archivo Excel
    df.to_excel(output_file, index=False)

    # Cargar el archivo Excel para aplicar formato
    wb = load_workbook(output_file)
    ws = wb.active

    # Aplicar formato de alineación y ajuste de tamaño de celdas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Aplicar color amarillo a la columna de calificaciones
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in ws["B"]:  # Columna B es la de calificaciones
        cell.fill = fill_yellow

    # Guardar cambios
    wb.save(output_file)
    print(f"Archivo Excel generado y formateado exitosamente: {output_file}")
    destroyInfo()