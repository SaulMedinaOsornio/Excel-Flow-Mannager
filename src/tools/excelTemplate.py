import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

#Variables finales
EVALUATION_TYPE = "ordinaria"
excelNameFile = 'ConcentradosCal.xlsx'
data = {
    'Matricula': [None],
    'Calificación': [None],
    'PeriodoEscolar': [None],
    'ClaveMateria': [None],
    'FechaAsig': [None],
    'Cuatrimestre': [None],
    'NivelEducativo': [None],
    'ClaveProfesor': [None],
    'Evaluación': [EVALUATION_TYPE]
}

def start():
    # Crear el DataFrame
    df = pd.DataFrame(data)

    # Reorganizar las columnas para que los nuevos campos estén primero
    column_order = [
        'Matricula', 'Calificación', 'PeriodoEscolar', 'ClaveMateria', 'FechaAsig',
        'Cuatrimestre', 'NivelEducativo', 'ClaveProfesor', 'Evaluación'
    ]

    df = df[column_order]

    # Guardar el DataFrame en un archivo de Excel
    df.to_excel(excelNameFile, index=False, engine='openpyxl')

    # Cargar el archivo de Excel
    wb = load_workbook(excelNameFile)
    ws = wb.active
    return ws, wb
def styleConfig(ws):
    # Definir el color amarillo
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Aplicar el color a todas las celdas de la columna "Calificación"
    for row in ws.iter_rows(min_row=1, max_row=30, min_col=2, max_col=2):
        for cell in row:
            cell.fill = yellow_fill

    # Ajustar el tamaño de las celdas al contenido del texto
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Obtener la letra de la columna
        for cell in column:
            try:
                # Calcular la longitud del contenido de la celda
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Ajustar el ancho de la columna
        adjusted_width = (max_length + 2)  # Añadir un pequeño margen
        ws.column_dimensions[column_letter].width = adjusted_width 
def close(wb):
    wb.save(excelNameFile)
    wb.close()
    
ws, wb = start()
styleConfig(ws)
close(wb)